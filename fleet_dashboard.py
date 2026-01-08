import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl.styles import PatternFill, Font
from openpyxl import load_workbook

# ---------------------------------------------------------
# 🔹 EXCEL STYLING + EXPORT FUNCTION
# ---------------------------------------------------------
def style_and_export_to_excel(df, sheet_name="Summary"):
    output = BytesIO()
    df.to_excel(output, index=False, sheet_name=sheet_name)

    wb = load_workbook(output)
    ws = wb.active

    # Header styling
    header_fill = PatternFill(start_color="007acc", end_color="007acc", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Bold subtotal rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if row[1].value == "Subtotal" or row[1].value == "Grand Total":
            for cell in row:
                cell.font = Font(bold=True)

    # Format ₦ columns
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '"₦"#,##0.00'

    final_output = BytesIO()
    wb.save(final_output)
    return final_output.getvalue()



# ---------------------------------------------------------
# 🔹 COMBINED SUPER-MODULE (Uploads CSV + Excel)
# ---------------------------------------------------------
def process_uploaded_files(uploaded_files):
    REQUIRED_COLS = ["Date", "Fleet", "Amount"]
    all_dfs = []

    progress = st.progress(0)
    status_text = st.empty()

    for i, file in enumerate(uploaded_files):
        file_name = file.name.lower()

        status_text.write(f"📄 Reading file: **{file.name}** ...")

        # CSV
        if file_name.endswith(".csv"):
            df = pd.read_csv(file)

        # Excel (auto sheet detection)
        elif file_name.endswith(".xlsx"):
            try:
                excel_file = pd.ExcelFile(file)
                sheet = excel_file.sheet_names[0]
                df = pd.read_excel(file, sheet_name=sheet)
            except:
                st.error(f"❌ Could not read sheet in `{file.name}`")
                continue

        else:
            st.warning(f"⚠️ Skipped unsupported file: {file.name}")
            continue

        # Validate Columns
        missing = [c for c in REQUIRED_COLS if c not in df.columns]
        if missing:
            st.error(f"❌ `{file.name}` missing columns: {missing}")
            continue

        # Standardize Format
        df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
        df["OnlyDate"] = df["Date"].dt.date
        df["Amount"] = pd.to_numeric(df["Amount"], errors="coerce").fillna(0)
        df["Fleet"] = df["Fleet"].astype(str)

        all_dfs.append(df)

        progress.progress((i + 1) / len(uploaded_files))

    status_text.write("✅ All files processed.")
    return all_dfs



# ---------------------------------------------------------
# 🔹 STREAMLIT APP UI — UNIFIED FUNCTION
# ---------------------------------------------------------
st.header("📊 Fleet Summary & Subtotals — Unified Processor")

uploaded_files = st.file_uploader(
    "Upload CSV or Excel files",
    type=["csv", "xlsx"],
    accept_multiple_files=True
)

if uploaded_files:
    dfs = process_uploaded_files(uploaded_files)

    if not dfs:
        st.stop()

    df = pd.concat(dfs, ignore_index=True)

    # ------------------ FILTER SECTION -------------------
    st.subheader("🔍 Filters")

    min_date, max_date = df["OnlyDate"].min(), df["OnlyDate"].max()
    date_range = st.date_input("Select Date Range", value=[min_date, max_date])

    if len(date_range) == 2:
        df = df[(df["OnlyDate"] >= date_range[0]) & (df["OnlyDate"] <= date_range[1])]

    fleet_filter = st.multiselect("Select Fleets", options=sorted(df["Fleet"].unique()))
    if fleet_filter:
        df = df[df["Fleet"].isin(fleet_filter)]
    # -------------------------------------------------------

    # ------------------ SUBTOTAL SECTION -------------------
    st.subheader("📌 Daily Subtotals")

    grouped = df.groupby(["OnlyDate", "Fleet"]).agg(
        TotalAmount=("Amount", "sum"),
        FleetCount=("Fleet", "count")
    ).reset_index()

    formatted_rows = []
    for date, group in grouped.groupby("OnlyDate"):
        sub_amt = group["TotalAmount"].sum()
        sub_cnt = group["FleetCount"].sum()

        for _, row in group.iterrows():
            formatted_rows.append({
                "Date": row["OnlyDate"],
                "Fleet": row["Fleet"],
                "Fleet Count": row["FleetCount"],
                "Total Amount (₦)": row["TotalAmount"]
            })

        # Subtotal Row
        formatted_rows.append({
            "Date": date,
            "Fleet": "Subtotal",
            "Fleet Count": sub_cnt,
            "Total Amount (₦)": sub_amt
        })

    subtotal_df = pd.DataFrame(formatted_rows)
    st.dataframe(subtotal_df)

    # ------------------ FLEET SUMMARY -------------------
    st.subheader("🚛 Fleet Summary (Combined Totals)")

    fleet_summary = df.groupby("Fleet").agg(
        TotalFleetCount=("Fleet", "count"),
        TotalAmount=("Amount", "sum")
    ).reset_index()

    # Add Grand Total Row
    fleet_summary.loc[len(fleet_summary)] = [
        "Grand Total",
        fleet_summary["TotalFleetCount"].sum(),
        fleet_summary["TotalAmount"].sum()
    ]

    st.dataframe(fleet_summary)

    # ------------------ DOWNLOADS -------------------
    st.subheader("⬇️ Export Results")

    styled_subtotal = style_and_export_to_excel(subtotal_df, "Daily_Subtotals")
    styled_fleet = style_and_export_to_excel(fleet_summary, "Fleet_Summary")

    st.download_button(
        "📥 Download Daily Subtotal Excel",
        styled_subtotal,
        "Daily_Subtotals.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    st.download_button(
        "📥 Download Fleet Summary Excel",
        styled_fleet,
        "Fleet_Summary.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )